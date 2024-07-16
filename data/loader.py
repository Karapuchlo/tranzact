import json
import csv
import openpyxl

def load_transactions_from_json(file_path):
    """
    Функция, которая загружает транзакции из JSON-файла.
    """
    with open(file_path, "r") as f:
        data = json.load(f)
        return data

def load_transactions_from_csv(file_path):
    """
    Функция, которая загружает транзакции из CSV-файла.
    """
    transactions = []
    with open(file_path, "r") as f:
        reader = csv.DictReader(f)
        for row in reader:
            transactions.append(row)
    return transactions


def load_transactions_from_xlsx(file_path):
    """
    Функция, которая загружает транзакции из XLSX-файла.
    """
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    transactions = []
    for row in range(2, sheet.max_row + 1):
        transaction = {
            "date": sheet.cell(row=row, column=1).value,
            "description": sheet.cell(row=row, column=2).value,
            "amount": sheet.cell(row=row, column=3).value,
            "currency": sheet.cell(row=row, column=4).value,
            "account": sheet.cell(row=row, column=5).value,
            "to": sheet.cell(row=row, column=6).value,
            "state": sheet.cell(row=row, column=7).value
        }
        transactions.append(transaction)
    return transactions

